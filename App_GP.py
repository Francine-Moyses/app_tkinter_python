from tkinter import *
from PIL import ImageTk, Image
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import ttk
import pandas as pd
from openpyxl import load_workbook
import datetime
import os
from tkinter import filedialog
import os.path
import shutil

######################
# ----INICIANDO APP----#
#######################
root = tk.Tk()
root.title("Gestão de Arquivos")
root.geometry('680x415')
root.configure(bg="#fff")
root.resizable(False, False)

#######################
# -----TELA LOGIN------#
#######################

# LOGO
canv = Canvas(root, width=0, height=0, bg='white')
canv.grid(row=2, column=3)
img = ImageTk.PhotoImage(file="/")
canv.create_image(20, 20, anchor=NW, image=img)
Label(root, image=img, bg='white').place(x=0, y=0)
frame = Frame(root, width=250, height=350, bg="white")
frame.place(x=350, y=10)

# LABEL LOGIN
heading = Label(frame, text='Login', fg='#4e8008', bg='white', font=('Microsoft YaHei UI Light', 23, 'bold'))
heading.place(x=90, y=15)

# ENTRADA USUÁRIO
user = Entry(frame, width=25, fg='black', border=0, bg='white', font=('Microsoft YaHei UI Light', 10))
user.place(x=30, y=100)
label1 = Label(frame, text="Nome de usuário", fg="gray", bg='white', font=('Microsoft YaHei UI Light', 7)).place(x=25,
                                                                                                                 y=127)
Frame(frame, width=295, height=1, bg='black').place(x=25, y=127)

# ENTRADA SENHA
code = Entry(frame, width=25, fg='black', border=0, bg='white', show='*', font=('Microsoft YaHei UI Light', 10))
code.place(x=30, y=170)
code.insert(0, '')
label2 = Label(frame, text="Digite sua senha", fg="gray", bg='white', font=('Microsoft YaHei UI Light', 7)).place(x=25,
                                                                                                                  y=197)
Frame(frame, width=295, height=1, bg='black').place(x=25, y=197)


#######################
# ----FUNÇÃO LOGIN-----#
#######################

def login():
    username = user.get()
    password = code.get()

    if username == 'admin' and password == 'admin':
        tela()
    elif username == '' and password == '':
        messagebox.showwarning("Gestão de Arquivos", "Digite usuário e senha")
    elif username == '':
        messagebox.showwarning("Gestão de Arquivos", "Necessário nome de usuário")
    elif password == '':
        messagebox.showwarning("Gestão de Arquivos", "Necessário senha")
    else:
        messagebox.showerror("Gestão de Arquivos", "Usuário e/ou senha inválidos")


def tela():
    root.destroy()
    secondWindow = tk.Tk()
    secondWindow.title("Gestão de Arquivos")
    secondWindow.configure(background="#ffffff")
    secondWindow.geometry("900x500")
    secondWindow.resizable(True, True)
    secondWindow.minsize(width=100, height=100)
    secondWindow.maxsize(width=1400, height=700)

    secondWindow.frame_1 = Frame(secondWindow, bg='#ffffff', highlightbackground='#3b5534',
                                 highlightthickness=1)  # highlightbackground = cor da borda do frame, highlightthickness cor da borda do frame
    secondWindow.frame_1.place(relx=0.21, rely=0.00, relwidth=0.78,
                               relheight=0.4)  # aumento e posições de elementos proporcionais, posicioes de 0 a 1. ex: x=0.30 = 30%, y=0.35 = 35%

    secondWindow.frame_2 = Frame(secondWindow, bg='#ffffff', highlightbackground='#3b5534', highlightthickness=1)
    secondWindow.frame_2.place(relx=0.21, rely=0.43, relwidth=0.78, relheight=0.56)

    secondWindow.frame_3 = Frame(secondWindow, bg='#3b5534', highlightbackground='#469536', highlightthickness=0)
    secondWindow.frame_3.place(relx=0.00, rely=0.00, relwidth=0.2, relheight=1)

    # ATUALIZAR DADOS
    def AtualizarDados():
        import pandas as pd
        import os
        import glob

        Dir = "/"
        dirout = "/"
        out = []
        os.chdir(f'{Dir}')
        for file_ant_ant in glob.glob("*"):
            if file_ant_ant != 'Thumbs.db':
                print('file_ant_ant', file_ant_ant)
                os.chdir(f'{baseDir}/{file_ant_ant}')
                for file_ant in glob.glob("*"):
                    print('file_ant', file_ant)
                    os.chdir(f'{Dir}/{file_ant_ant}/{file_ant}')
                    for file in glob.glob("*"):
                        print('file', file)
                        if file != 'Thumbs.db':
                            os.chdir(f'{Dir}/{file_ant_ant}/{file_ant}/{file}')
                            for file_pos in glob.glob("*"):
                                print('file_pos', file_pos)
                                out.append([file_ant_ant, file_ant, file, file_pos])

        df_out = pd.DataFrame(data=out, columns=['PASTA1', 'PASTA2', 'PASTA3', 'ARQUIVO'])

        df_out.to_excel(dirout + '\', index=False)
        # print(dirout+ '\')

        filename1 = "/"

        df1 = pd.read_excel(filename1)

        # Arquivo 1
        filter_full = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_prog1 = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_prog2 = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_prog3 = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_prog4 = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_prog5 = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_prog6 = df1.query("ARQUIVO.str.contains('S001')")

        prog_1 = filter_full
        prog_1["ARQUIVO"] = filter_full["ARQUIVO"].str.startswith('001')

        prog_2 = filter_full_prog1
        prog_2["ARQUIVO"] = filter_full_prog1["ARQUIVO"].str.startswith('002')

        prog_3 = filter_full_prog2
        prog_3["ARQUIVO"] = filter_full_prog2["ARQUIVO"].str.startswith('003')

        prog_4 = filter_full_prog3
        prog_4["ARQUIVO"] = filter_full_prog3["ARQUIVO"].str.startswith('004')

        prog_5 = filter_full_prog4
        prog_5["ARQUIVO"] = filter_full_prog4["ARQUIVO"].str.startswith('005')

        prog_6 = filter_full_prog5
        prog_6["ARQUIVO"] = filter_full_prog5["ARQUIVO"].str.startswith('006')

        prog_7 = filter_full_prog6
        prog_7["ARQUIVO"] = filter_full_prog6["ARQUIVO"].str.startswith('007')

        frames = [prog_1, prog_2, prog_3, prog_4, prog_5, prog_6, prog_7]

        df_concf = pd.concat(frames)

        prog_ = df_concf[["PASTA1", "ARQUIVO"]]
        df_prog = prog_.rename(columns={'ARQUIVO': 'Programação'})
        df_prog['P_arquivo'] = df_prog['P_arquivo'].map({bool(True): 1},
                                                            na_action=None)

        df_progf = df_prog.query("P_arquivo == 1")
        df_progG = df_progf.groupby("PASTA1")["P_arquivo"].count()

        # Arquivo 2
        filter_full_op = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_op1 = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_op2 = df1.query("ARQUIVO.str.contains('S001')")

        op_1 = filter_full_op
        op_1["ARQUIVO"] = filter_full_op["ARQUIVO"].str.startswith('008')

        op_2 = filter_full_op1
        op_2["ARQUIVO"] = filter_full_op1["ARQUIVO"].str.startswith('009')

        op_3 = filter_full_op2
        op_3["ARQUIVO"] = filter_full_op2["ARQUIVO"].str.startswith('010')

        frames_O = [op_1, op_2, op_3]

        df_concf_O = pd.concat(frames_O)

        op_ = df_concf_O[["PASTA1", "ARQUIVO"]]
        df_op = op_.rename(columns={'ARQUIVO': 'Q_arquivo'})
        df_op['P_arquivo'] = df_op['Q_arquivo'].map({bool(True): 1},
                                                     na_action=None)

        df_opf = df_op.query("Q_arqyivo == 1")
        df_opG = df_opf.groupby("PASTA1")["Q_arqyuvo"].count()

        # Arquivo 3
        filter_full_nue = df1.query("ARQUIVO.str.contains('S001')")
        filter_full_nue1 = df1.query("ARQUIVO.str.contains('S001')")

        nue_1 = filter_full_nue
        nue_1["ARQUIVO"] = filter_full_nue["ARQUIVO"].str.startswith('011')

        nue_2 = filter_full_nue1
        nue_2["ARQUIVO"] = filter_full_nue1["ARQUIVO"].str.startswith('012')

        frames_N = [nue_1, nue_2]

        df_concf_N = pd.concat(frames_N)

        nue_ = df_concf_N[["PASTA1", "ARQUIVO"]]
        df_nue = nue_.rename(columns={'ARQUIVO': 'R_arquivo'})
        df_nue['P_arquivo'] = df_nue['R_arquivo'].map({bool(True): 1},
                                                         na_action=None)

        df_nuef = df_nue.query("R_arquivo == 1")
        df_nueG = df_nuef.groupby("PASTA1")["R_arqyivo"].count()

        df_merge = pd.merge(df_progG, df_opG, how='outer', on='PASTA1')
        df_dinamico = pd.merge(df_merge, df_nueG, how='outer', on='PASTA1')

        df_dinamico['P_arquivo'] = df_dinamico['P_arquivo'].fillna(0)
        df_dinamico['Operação'] = df_dinamico['Q_arquivo'].fillna(0)
        df_dinamico['Unitização'] = df_dinamico['R_arquivo'].fillna(0)

        dinamico = pd.DataFrame(data=df_dinamico)
        dinamico.to_excel('/')

        messagebox.showinfo("Atualizar Dados", "Dados atualizados com sucesso.")

    secondWindow.button_atualizar = Button(secondWindow.frame_3, text="Atualizar Dados", bg='#3b5534', fg='white',
                                           border=1, command=AtualizarDados)
    secondWindow.button_atualizar.place(relx=0.15, rely=0.82, relwidth=0.7, relheight=0.08)

    # TREEVIEW DADOS BRUTOS
    secondWindow.paned = ttk.PanedWindow(secondWindow.frame_2, width=250)
    secondWindow.paned.place(relx=0.01, rely=0.1, relwidth=0.99, relheight=0.85)

    secondWindow.pane_1 = ttk.Frame(secondWindow.paned, padding=5)
    secondWindow.paned.add(secondWindow.pane_1, weight=1)

    secondWindow.scrollbar = ttk.Scrollbar(secondWindow.pane_1)
    secondWindow.scrollbar.pack(side="right", fill="y")

    secondWindow.treeview = ttk.Treeview(
        secondWindow.pane_1,
        selectmode="browse",
        yscrollcommand=secondWindow.scrollbar.set,
        columns=(1, 2),
        height=10,
    )
    secondWindow.treeview.pack(expand=True, fill="both")
    secondWindow.scrollbar.config(command=secondWindow.treeview.yview)

    filename = "/"
    df = pd.read_excel(filename)

    columns = list(df.columns)
    secondWindow.treeview["columns"] = columns

    secondWindow.treeview.heading("#0", text="", anchor="w")
    secondWindow.treeview.heading("#1", text="EPS", anchor="w")
    secondWindow.treeview.heading("#2", text="Contrato/Ano", anchor="w")
    secondWindow.treeview.heading("#3", text="Pasta", anchor="w")
    secondWindow.treeview.heading("#4", text="Arquivo", anchor="w")

    secondWindow.treeview.column("#0", width=50)
    secondWindow.treeview.column("#1", width=90)
    secondWindow.treeview.column("#2", width=90)
    secondWindow.treeview.column("#3", width=90)
    secondWindow.treeview.column("#4", width=200)

    for index, row in df.iterrows():
        secondWindow.treeview.insert("", "end", text=index, values=list(row))

    def select_EPS(*args):
        secondWindow.treeview.delete(*secondWindow.treeview.get_children())
        for index, row in df.loc[df["PASTA1"].eq(secondWindow.optionemenu_EPS.get())].iterrows():
            secondWindow.treeview.insert("", "end", text=index, values=list(row))

    def select_C(*args):
        secondWindow.treeview.delete(*secondWindow.treeview.get_children())
        for index, row in df.loc[df["PASTA"].eq(secondWindow.optionemenu_CONTRATO.get())].iterrows():
            secondWindow.treeview.insert("", "end", text=index, values=list(row))

    def select_P(*args):
        pesproj = secondWindow.entry_p.get()
        proj = df["PASTA2"].unique()
        projD = df1["PASTA2"].unique()

        if pesproj in proj:
            secondWindow.treeview.delete(*secondWindow.treeview.get_children())
            comparedValue = secondWindow.entry_projeto.get()
            for index, row in df.loc[df["PASTA1"] == comparedValue].iterrows():
                secondWindow.treeview.insert("", "end", text=index, values=list(row))
        else:
            messagebox.showinfo("Gestão de Arquivos", "inexistente na Pasta")

        if pesp in pD:
            secondWindow.treeviewD.delete(*secondWindow.treeviewD.get_children())
            comparedValueD = secondWindow.entry_p.get()
            for index, row in df1.loc[df1["PASTA2"] == comparedValueD].iterrows():
                secondWindow.treeviewD.insert("", "end", text=index, values=list(row))

    def Limpar():
        secondWindow.treeview.delete(*secondWindow.treeview.get_children())
        for index, row in df.iterrows():
            secondWindow.treeview.insert("", "end", text=index, values=list(row))

        secondWindow.treeviewD.delete(*secondWindow.treeviewD.get_children())
        for index, row in df1.iterrows():
            secondWindow.treeviewD.insert("", "end", text=index, values=list(row))

    secondWindow.treeview.pack()

    # TREEVIEW TABELAD
    secondWindow.panedD = ttk.PanedWindow(secondWindow.frame_1, width=250)
    secondWindow.panedD.place(relx=0.01, rely=0.1, relwidth=0.99, relheight=0.85)

    secondWindow.pane_1D = ttk.Frame(secondWindow.panedD, padding=5)
    secondWindow.panedD.add(secondWindow.pane_1D, weight=1)

    secondWindow.scrollbarD = ttk.Scrollbar(secondWindow.pane_1D)
    secondWindow.scrollbarD.pack(side="right", fill="y")

    secondWindow.treeviewD = ttk.Treeview(
        secondWindow.pane_1D,
        selectmode="browse",
        yscrollcommand=secondWindow.scrollbarD.set,
        columns=(1, 2),
        height=10,
    )
    secondWindow.treeviewD.pack(expand=True, fill="both")
    secondWindow.scrollbarD.config(command=secondWindow.treeviewD.yview)

    filename1 = "/"
    df_1 = pd.read_excel(filename1)
    df1 = df_1[["PASTA2", "P_arquivo", "Q_arquivo", "R_arquivo"]]

    columns = list(df1.columns)
    secondWindow.treeviewD["columns"] = columns

    secondWindow.treeviewD.heading("#0", text="", anchor="w")
    secondWindow.treeviewD.heading("#1", text="Pasta", anchor="w")
    secondWindow.treeviewD.heading("#2", text="P_arquivo")
    secondWindow.treeviewD.heading("#3", text="Q_arquivo")
    secondWindow.treeviewD.heading("#4", text="R_arquivo")

    secondWindow.treeviewD.column("#0", width=30)
    secondWindow.treeviewD.column("#1", width=100)
    secondWindow.treeviewD.column("#2", width=120, anchor="c")
    secondWindow.treeviewD.column("#3", width=120, anchor="c")
    secondWindow.treeviewD.column("#4", width=120, anchor="c")

    for index, row in df1.iterrows():
        secondWindow.treeviewD.insert("", "end", text=index, values=list(row))

    secondWindow.treeview.pack()

    #######################
    # -----FILTROS--------#
    #######################

    # FILTRO
    secondWindow.label_projeto = Label(secondWindow.frame_3, text="BEM-VINDO", anchor="w", bg='#3b5534', fg='#ffffff',
                                       font=('Microsoft YaHei UI Light', 16))
    secondWindow.label_projeto.place(relx=0.06, rely=0.01, relwidth=0.9, relheight=0.2)

    secondWindow.label_data = Label(secondWindow.frame_3, text=f"{datetime.date.today(): %d/%m/%y}", anchor="w",
                                    bg='#3b5534', fg='#ffffff', font=('Microsoft YaHei UI Light', 8))
    secondWindow.label_data.place(relx=0.68, rely=0.01, relwidth=0.4, relheight=0.04)

    # PESQUISAR
    secondWindow.label_projeto = Label(secondWindow.frame_3, text="ARQUIVOS", anchor="w", bg='#3b5534',
                                       fg='#ffffff', font=('Microsoft YaHei UI Light', 7))
    secondWindow.label_projeto.place(relx=0.06, rely=0.2, relwidth=0.6, relheight=0.04)

    secondWindow.entry_projeto = Entry(secondWindow.frame_3, bg='#ffffff', fg='#3b5534', border=0)
    secondWindow.entry_projeto.place(relx=0.06, rely=0.25, relwidth=0.8, relheight=0.04)

    secondWindow.button_projeto = Button(secondWindow.frame_3, text='pesquisar', bg='#3b5534', fg='white', border=0,
                                         command=select_P)
    secondWindow.button_projeto.place(relx=0.46, rely=0.3, relwidth=0.4, relheight=0.04)

    secondWindow.button_p = Button(secondWindow.frame_2, text='limpar', bg='#ffffff', fg='#3b5534', border=0,
                                         command=Limpar)
    secondWindow.button_p.place(relx=0.85, rely=0.01, relwidth=0.150, relheight=0.07)

    secondWindow.label_treeview = Label(secondWindow.frame_2, text='ARQUIVOS', bg='#ffffff', fg='#3b5534',
                                        border=0)
    secondWindow.label_treeview.place(relx=0.35, rely=0.02, relwidth=0.25, relheight=0.07)

    secondWindow.button_projeto = Button(secondWindow.frame_1, text='limpar', bg='#ffffff', fg='#3b5534', border=0,
                                         command=Limpar)
    secondWindow.button_projeto.place(relx=0.85, rely=0.01, relwidth=0.150, relheight=0.07)

    secondWindow.label_treeview = Label(secondWindow.frame_1,""', bg='#ffffff', fg='#3b5534',
                                        border=0)
    secondWindow.label_treeview.place(relx=0.35, rely=0.02, relwidth=0.25, relheight=0.07)

    # EXPORTAR
    def ExportarDados():
        arquivo = pd.DataFrame(df)
        dinamico = pd.DataFrame(df1)

        writer = pd.ExcelWriter('/arquivoGP.xlsx',
                                engine='xlsxwriter')

        arquivo.to_excel(writer, sheet_name='PASTA')
        dinamico.to_excel(writer, sheet_name='Pastas')
        writer.save()

        messagebox.showinfo("Gestão de Arquivos", "Planilha salva no diretório. Nome: arquivoGP.xlsx")

    secondWindow.button_atualizar = Button(secondWindow.frame_3, text="Exportar Dados", bg='#3b5534', fg='white',
                                           border=1, command=ExportarDados)
    secondWindow.button_atualizar.place(relx=0.15, rely=0.905, relwidth=0.7, relheight=0.08)

    # CRIAR PASTA
    secondWindow.label_projeto = Label(secondWindow.frame_3, text="CRIAR PASTA", anchor="w", bg='#3b5534', fg='#ffffff',
                                       font=('Microsoft YaHei UI Light', 7))
    secondWindow.label_projeto.place(relx=0.15, rely=0.43, relwidth=0.7, relheight=0.05)

    # MANIPULAR ARQUIVOS
    secondWindow.label_projeto = Label(secondWindow.frame_3, text="MANIPULAR ARQUIVOS", anchor="w", bg='#3b5534',
                                       fg='#ffffff', font=('Microsoft YaHei UI Light', 7))
    secondWindow.label_projeto.place(relx=0.15, rely=0.6, relwidth=0.7, relheight=0.05)

    # CRIAR PASTA
    def CriarPasta():
        diretório = '/'
        arquivo = filedialog.askopenfilename()
        df_cpp = pd.read_excel(arquivo)
        df_cpp1 = df_cpp["PASTA1"].count() + 1
        linhas = df_cpp1
        wb = load_workbook(arquivo)
        ws = wb.active
        n = linhas
        for row in range(2, n):
            PASTA_ANT = ws.cell(row=row, column=1).value
            PASTA = ws.cell(row=row, column=2).value
            PASTA_1 = ws.cell(row=row, column=3).value
            PASTA_ANT1 = str(PASTA_ANT)
            PASTA1 = str(PASTA)
            PASTA_11 = str(PASTA_1)
            if not os.path.exists(diretório + PASTA_ANT1 + '/' + PASTA1 + '/' + PASTA_11):
                os.makedirs(diretório + PASTA_ANT1 + '/' + PASTA1 + '/' + PASTA_11)

        messagebox.showinfo("Gestão de Arquivos", "Pastas Criadas")

    secondWindow.button_atualizar = Button(secondWindow.frame_3, text="Nome", bg='#3b5534', fg='white',
                                           border=1, "criar pasta")
    secondWindow.button_atualizar.place(relx=0.15, rely=0.475, relwidth=0.7, relheight=0.05)

    # CRIAR PASTA
    def CriarPastaProjetos():
        diretório = '/'
        arquivo = filedialog.askopenfilename()
        df_cp = pd.read_excel(arquivo)
        df_cp1 = df_cp["PASTA1"].count() + 1
        linhas = df_cp1
        wb = load_workbook(arquivo)
        ws = wb.active
        n = linhas
        for row in range(2, n):
            PASTA1 = ws.cell(row=row, column=1).value
            PASTA = ws.cell(row=row, column=2).value
            PASTA_1 = ws.cell(row=row, column=3).value
            PASTA1 = str(PASTA_ANT)
            PASTA1 = str(PASTA)
            PASTA_11 = str(PASTA_1)
            if not os.path.exists(diretório + PASTA1 + '/' + PASTA1 + '/' + PASTA_11):
                os.makedirs(diretório + PASTA1 + '/' + PASTA1 + '/' + PASTA_11)

        messagebox.showinfo("Gestão de arquivos", "Pastas criadas")

    secondWindow.button_atualizar = Button(secondWindow.frame_3, text="Projetos", bg='#3b5534', fg='white', border=1,
                                           command=CriarPasta)
    secondWindow.button_atualizar.place(relx=0.15, rely=0.525, relwidth=0.7, relheight=0.05)

    # RENOMEAR ARQUIVOS
    def RenomearArquivos():
        diretório = '/'
        arquivo = filedialog.askopenfilename()
        df = pd.read_excel(arquivo)
        df1 = df["PASTA_ANT"].count() + 1
        linhas = df1
        wb = load_workbook(arquivo)
        ws = wb.active
        n = linhas

        for row in range(2, n):
            PASTA_ANT = ws.cell(row=row, column=1).value
            PASTA = ws.cell(row=row, column=2).value
            PASTA_1 = ws.cell(row=row, column=3).value
            N_DOC = ws.cell(row=row, column=4).value
            RENOME_DOC = ws.cell(row=row, column=5).value
            PASTA_ANT1 = str(PASTA_ANT)
            PASTA1 = str(PASTA)
            PASTA_11 = str(PASTA_1)
            N_DOC1 = str(N_DOC)
            RENOME_DOC1 = str(RENOME_DOC)
            if not os.path.exists(diretório + PASTA_ANT1 + '/' + PASTA1 + '/' + PASTA_11 + '/' + RENOME_DOC1):
                os.rename(diretório + PASTA_ANT1 + '/' + PASTA1 + '/' + PASTA_11 + '/' + N_DOC1,
                          diretório + PASTA_ANT1 + '/' + PASTA1 + '/' + PASTA_11 + '/' + RENOME_DOC1)
        messagebox.showinfo("Gestão de arquivos", "Arquivos renomeados")

    secondWindow.button_atualizar = Button(secondWindow.frame_3, text="Renomear", bg='#3b5534', fg='white', border=1,
                                           command=RenomearArquivos)
    secondWindow.button_atualizar.place(relx=0.15, rely=0.645, relwidth=0.7, relheight=0.05)

    # MOVER ARQUIVOS
    def MoverArquivos():
        diretório = '/'
        arquivo = filedialog.askopenfilename()
        df = pd.read_excel(arquivo)
        df1 = df["PASTA_ANT"].count() + 1
        linhas = df1
        wb = load_workbook(arquivo)
        ws = wb.active
        n = linhas

        for row in range(2, n):
            PASTA_ANT = ws.cell(row=row, column=1).value
            PASTA = ws.cell(row=row, column=2).value
            PASTA1 = ws.cell(row=row, column=3).value
            ARQUIVO = ws.cell(row=row, column=4).value
            DIRETORIO = ws.cell(row=row, column=5).value
            PASTA_ANT1 = str(PASTA_ANT)
            PASTA1 = str(PASTA)
            PASTA_11 = str(PASTA1)
            ARQUIVO1 = str(ARQUIVO)
            DIRETORIO1 = str(DIRETORIO)
            shutil.move(DIRETORIO1 + ARQUIVO1 + '.xlsx', diretório + PASTA_ANT1 + '/' + PASTA1 + '/' + PASTA_11)
        messagebox.showinfo("Gestão de arquivos", "Arquivos movidos")

    secondWindow.button_atualizar = Button(secondWindow.frame_3, text="Mover", bg='#3b5534', fg='white', border=1,
                                           command=MoverArquivos)
    secondWindow.button_atualizar.place(relx=0.15, rely=0.695, relwidth=0.7, relheight=0.05)


# BOTÃO DE ACESSO
Button(frame, width=25, pady=7, text='Acesse', bg='#4e8008', fg='white', border=0, command=login).place(x=50, y=280)

root.mainloop()
